"""Parser for age-stratified population data from Excel files"""
import openpyxl
from typing import Dict, List, Optional
from pathlib import Path


class AgeGroupParser:
    """Parse age-stratified population data (年齢階級別人口) from Excel files"""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.worksheet = None

    def load(self):
        """Load the Excel file"""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Age group population data file not found: {self.file_path}")

        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        # Sheet name: 年齢別人口(市区町村別)【総計】
        self.worksheet = self.workbook[self.workbook.sheetnames[0]]

    def parse(self) -> List[Dict]:
        """
        Parse age-stratified population data from Excel file

        Returns:
            List of dictionaries with age group data for each municipality
            Each municipality has 3 entries: 計 (total), 男 (male), 女 (female)
        """
        if self.worksheet is None:
            self.load()

        ws = self.worksheet
        data = []

        # Data starts at row 4
        # Columns:
        # A(1): 団体コード, B(2): 都道府県名, C(3): 市区町村名, D(4): 性別
        # E(5): 総数
        # F-Z(6-26): Age groups (0-4歳 to 100歳以上)
        age_group_names = [
            "0-4歳", "5-9歳", "10-14歳", "15-19歳", "20-24歳",
            "25-29歳", "30-34歳", "35-39歳", "40-44歳", "45-49歳",
            "50-54歳", "55-59歳", "60-64歳", "65-69歳", "70-74歳",
            "75-79歳", "80-84歳", "85-89歳", "90-94歳", "95-99歳",
            "100歳以上"
        ]

        for row_idx in range(4, ws.max_row + 1):
            jichitai_code = ws.cell(row_idx, 1).value
            prefecture = ws.cell(row_idx, 2).value
            municipality = ws.cell(row_idx, 3).value
            gender = ws.cell(row_idx, 4).value

            # Skip if no code or municipality name
            if not jichitai_code or not municipality:
                continue

            # Skip if code is not a proper 6-digit code
            code_str = str(jichitai_code).strip()
            if not code_str.isdigit() or len(code_str) != 6:
                continue

            # Get total population
            total = ws.cell(row_idx, 5).value

            # Helper to safely convert to int
            def safe_int(val):
                if val is None:
                    return None
                try:
                    return int(val)
                except (ValueError, TypeError):
                    return None

            # Parse age groups (columns 6-26)
            age_groups = {}
            for i, age_group_name in enumerate(age_group_names):
                col_idx = 6 + i  # Start from column F (6)
                value = ws.cell(row_idx, col_idx).value
                age_groups[age_group_name] = safe_int(value)

            record = {
                "jichitai_code": code_str.zfill(6),
                "prefecture": prefecture,
                "municipality": municipality,
                "gender": gender,  # 計, 男, or 女
                "total": safe_int(total),
                "age_groups": age_groups
            }

            data.append(record)

        return data

    def get_by_code(self, jichitai_code: str) -> List[Dict]:
        """
        Get age-stratified population data for a specific municipality by code

        Returns:
            List of 3 records (計, 男, 女) for the municipality
        """
        data = self.parse()
        code = str(jichitai_code).zfill(6)

        results = []
        for record in data:
            if record["jichitai_code"] == code:
                results.append(record)

        return results

    def get_by_name(self, municipality_name: str, prefecture: Optional[str] = None) -> List[Dict]:
        """
        Get age-stratified population data for municipalities by name

        Args:
            municipality_name: Name of municipality to search
            prefecture: Optional prefecture name to narrow search

        Returns:
            List of matching records (3 per municipality: 計, 男, 女)
        """
        data = self.parse()
        results = []

        for record in data:
            if municipality_name in record["municipality"]:
                if prefecture is None or prefecture in record["prefecture"]:
                    results.append(record)

        return results

    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()
