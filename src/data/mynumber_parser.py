"""Parser for My Number Card issuance rate data from MIC"""
from typing import Dict, List, Optional
from pathlib import Path
import openpyxl


class MyNumberParser:
    """Parser for My Number Card issuance rate data"""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.wb = None
        self.ws = None

        if self.file_path.exists():
            self.wb = openpyxl.load_workbook(str(self.file_path), data_only=True)
            # シート名は「公表用」
            if "公表用" in self.wb.sheetnames:
                self.ws = self.wb["公表用"]
            else:
                self.ws = self.wb.active

    def parse(self) -> List[Dict]:
        """
        Parse the My Number Card data file

        Returns:
            List of dictionaries containing municipality data
        """
        if not self.ws:
            return []

        results = []

        # データは119行目から開始（118行目は全国集計）
        for row_idx in range(119, self.ws.max_row + 1):
            prefecture = self.ws.cell(row_idx, 1).value  # A列: 都道府県名
            municipality = self.ws.cell(row_idx, 2).value  # B列: 市区町村名
            population = self.ws.cell(row_idx, 3).value  # C列: 人口
            issued_cards = self.ws.cell(row_idx, 4).value  # D列: 保有枚数
            issuance_rate = self.ws.cell(row_idx, 5).value  # E列: 交付率

            # 空行をスキップ
            if not prefecture or not municipality:
                continue

            # Convert issuance rate to percentage (data is in decimal form like 0.78 = 78%)
            rate = self._safe_float(issuance_rate)
            if rate is not None:
                rate = rate * 100  # Convert to percentage

            record = {
                "prefecture": str(prefecture).strip() if prefecture else None,
                "municipality": str(municipality).strip() if municipality else None,
                "mynumber_card": {
                    "population": self._safe_int(population),
                    "issued_cards": self._safe_int(issued_cards),
                    "issuance_rate": round(rate, 2) if rate is not None else None,
                }
            }

            results.append(record)

        return results

    def get_by_name(self, municipality_name: str, prefecture: Optional[str] = None) -> Optional[Dict]:
        """
        Get My Number Card data by municipality name

        Args:
            municipality_name: Municipality name
            prefecture: Prefecture name (optional, for disambiguation)

        Returns:
            Dictionary with My Number Card data
        """
        all_data = self.parse()

        for record in all_data:
            if record["municipality"] == municipality_name:
                if prefecture is None or record["prefecture"] == prefecture:
                    return record

        return None

    def _safe_int(self, val) -> Optional[int]:
        """Safely convert value to int"""
        if val is None:
            return None
        try:
            return int(val)
        except (ValueError, TypeError):
            return None

    def _safe_float(self, val) -> Optional[float]:
        """Safely convert value to float"""
        if val is None:
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    def close(self):
        """Close the workbook"""
        if self.wb:
            self.wb.close()