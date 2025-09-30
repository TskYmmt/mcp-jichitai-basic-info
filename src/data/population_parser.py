"""Parser for population data from Excel files"""
import openpyxl
from typing import Dict, List, Optional
from pathlib import Path


class PopulationParser:
    """Parse municipal population data from Excel files"""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.worksheet = None

    def load(self):
        """Load the Excel file"""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Population data file not found: {self.file_path}")

        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        # Get first sheet (人口、世帯数、人口動態（市区町村別）【総計】)
        self.worksheet = self.workbook[self.workbook.sheetnames[0]]

    def parse(self) -> List[Dict]:
        """
        Parse population data from Excel file

        Returns:
            List of dictionaries with population data for each municipality
        """
        if self.worksheet is None:
            self.load()

        ws = self.worksheet
        data = []

        # Data starts at row 7 (全国合計), municipalities start at row 9
        # Columns: 1=団体コード, 2=都道府県名, 3=市区町村名, 4=人口(男), 5=人口(女), 6=人口(計), 7=世帯数
        for row_idx in range(9, ws.max_row + 1):
            jichitai_code = ws.cell(row_idx, 1).value
            prefecture = ws.cell(row_idx, 2).value
            municipality = ws.cell(row_idx, 3).value

            # Skip if no code or municipality name is "-" (prefecture summary row)
            if not jichitai_code or municipality == "-":
                continue

            # Skip if code is not a proper 6-digit code
            if not str(jichitai_code).isdigit() or len(str(jichitai_code)) != 6:
                continue

            pop_male = ws.cell(row_idx, 4).value
            pop_female = ws.cell(row_idx, 5).value
            pop_total = ws.cell(row_idx, 6).value
            households = ws.cell(row_idx, 7).value

            # Additional population dynamics data
            transfer_in_domestic = ws.cell(row_idx, 8).value
            transfer_in_foreign = ws.cell(row_idx, 9).value
            transfer_in_total = ws.cell(row_idx, 10).value
            births = ws.cell(row_idx, 11).value

            # Helper to safely convert to int
            def safe_int(val):
                if val is None:
                    return None
                try:
                    return int(val)
                except (ValueError, TypeError):
                    return None

            record = {
                "jichitai_code": str(jichitai_code).zfill(6),
                "prefecture": prefecture,
                "municipality": municipality,
                "population": {
                    "total": safe_int(pop_total),
                    "male": safe_int(pop_male),
                    "female": safe_int(pop_female),
                },
                "households": safe_int(households),
                "population_dynamics": {
                    "transfer_in_domestic": safe_int(transfer_in_domestic),
                    "transfer_in_foreign": safe_int(transfer_in_foreign),
                    "transfer_in_total": safe_int(transfer_in_total),
                    "births": safe_int(births),
                }
            }

            data.append(record)

        return data

    def get_by_code(self, jichitai_code: str) -> Optional[Dict]:
        """Get population data for a specific municipality by code"""
        data = self.parse()
        code = str(jichitai_code).zfill(6)

        for record in data:
            if record["jichitai_code"] == code:
                return record

        return None

    def get_by_name(self, municipality_name: str, prefecture: Optional[str] = None) -> List[Dict]:
        """
        Get population data for municipalities by name

        Args:
            municipality_name: Name of municipality to search
            prefecture: Optional prefecture name to narrow search

        Returns:
            List of matching records
        """
        data = self.parse()
        results = []

        for record in data:
            if municipality_name in record["municipality"]:
                if prefecture is None or prefecture in record["prefecture"]:
                    results.append(record)

        return results

    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()