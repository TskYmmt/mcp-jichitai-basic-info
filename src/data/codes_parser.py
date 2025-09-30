"""Parser for municipal codes from Excel files"""
import openpyxl
from typing import Dict, List, Optional
from pathlib import Path


class CodesParser:
    """Parse municipal codes from Excel files"""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.worksheet = None

    def load(self):
        """Load the Excel file"""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Codes data file not found: {self.file_path}")

        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        # Get first sheet (R6.1.1現在の団体)
        self.worksheet = self.workbook[self.workbook.sheetnames[0]]

    def parse(self) -> List[Dict]:
        """
        Parse municipal codes from Excel file

        Returns:
            List of dictionaries with code data for each municipality
        """
        if self.worksheet is None:
            self.load()

        ws = self.worksheet
        data = []

        # Data starts at row 2
        # Columns: 1=団体コード, 2=都道府県名(漢字), 3=市区町村名(漢字), 4=都道府県名(カナ), 5=市区町村名(カナ)
        for row_idx in range(2, ws.max_row + 1):
            jichitai_code = ws.cell(row_idx, 1).value
            prefecture_kanji = ws.cell(row_idx, 2).value
            municipality_kanji = ws.cell(row_idx, 3).value
            prefecture_kana = ws.cell(row_idx, 4).value
            municipality_kana = ws.cell(row_idx, 5).value

            # Skip if no code
            if not jichitai_code:
                continue

            # Determine type
            jichitai_type = None
            if municipality_kanji is None:
                jichitai_type = "都道府県"
            elif "市" in municipality_kanji:
                if "区" in municipality_kanji:
                    jichitai_type = "区"
                else:
                    jichitai_type = "市"
            elif "町" in municipality_kanji:
                jichitai_type = "町"
            elif "村" in municipality_kanji:
                jichitai_type = "村"

            record = {
                "jichitai_code": str(jichitai_code).zfill(6),
                "prefecture": prefecture_kanji,
                "municipality": municipality_kanji,
                "prefecture_kana": prefecture_kana,
                "municipality_kana": municipality_kana,
                "jichitai_type": jichitai_type,
            }

            data.append(record)

        return data

    def get_by_code(self, jichitai_code: str) -> Optional[Dict]:
        """Get code data for a specific municipality by code"""
        data = self.parse()
        code = str(jichitai_code).zfill(6)

        for record in data:
            if record["jichitai_code"] == code:
                return record

        return None

    def get_by_name(self, municipality_name: str, prefecture: Optional[str] = None) -> List[Dict]:
        """
        Get code data for municipalities by name (fuzzy match)

        Args:
            municipality_name: Name of municipality to search
            prefecture: Optional prefecture name to narrow search

        Returns:
            List of matching records with match scores
        """
        data = self.parse()
        results = []

        for record in data:
            # Skip prefecture-only records
            if record["municipality"] is None:
                continue

            # Check if name matches
            municipality = record["municipality"]
            match_score = 0.0

            if municipality_name == municipality:
                match_score = 1.0  # Exact match
            elif municipality_name in municipality:
                match_score = 0.9  # Contains
            elif municipality in municipality_name:
                match_score = 0.8  # Partial match

            # Check prefecture if specified
            if match_score > 0:
                if prefecture is None or prefecture in record["prefecture"]:
                    result = record.copy()
                    result["match_score"] = match_score
                    results.append(result)

        # Sort by match score descending
        results.sort(key=lambda x: x["match_score"], reverse=True)
        return results

    def get_municipalities_only(self) -> List[Dict]:
        """Get only municipality records (exclude prefectures)"""
        data = self.parse()
        return [r for r in data if r["municipality"] is not None]

    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()