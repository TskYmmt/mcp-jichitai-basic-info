"""Parser for finance data from Excel files"""
import openpyxl
from typing import Dict, List, Optional
from pathlib import Path


class FinanceParser:
    """Parse municipal finance data from Excel files"""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.worksheet = None

    def load(self):
        """Load the Excel file"""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Finance data file not found: {self.file_path}")

        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        # Get first sheet
        self.worksheet = self.workbook[self.workbook.sheetnames[0]]

    def parse(self) -> List[Dict]:
        """
        Parse finance summary data from Excel file

        Data source: 全市町村の主要財政指標（令和5年度）
        Coverage: All municipalities (cities, towns, villages, special wards)

        Returns:
            List of dictionaries with finance data for each municipality
        """
        if self.worksheet is None:
            self.load()

        ws = self.worksheet
        data = []

        # Data starts at row 3 (row 2 is header)
        # Key columns:
        # 1: 団体コード
        # 2: 都道府県名
        # 3: 団体名
        # 4: 財政力指数
        # 5: 経常収支比率
        # 6: 実質公債費比率
        # 7: 将来負担比率
        # 8: ラスパイレス指数

        # Helper to convert "-" to None
        def normalize_value(val):
            if val == "-" or val == "－":
                return None
            return val

        for row_idx in range(3, ws.max_row + 1):
            jichitai_code = ws.cell(row_idx, 1).value

            # Skip if no code
            if not jichitai_code or not str(jichitai_code).strip():
                continue

            # Skip if not a valid 6-digit code
            if not str(jichitai_code).isdigit() or len(str(jichitai_code)) != 6:
                continue

            prefecture_name = ws.cell(row_idx, 2).value
            municipality_name = ws.cell(row_idx, 3).value
            financial_capability_index = ws.cell(row_idx, 4).value
            current_balance_ratio = ws.cell(row_idx, 5).value
            real_debt_service_ratio = ws.cell(row_idx, 6).value
            future_burden_ratio = ws.cell(row_idx, 7).value
            laspeyres_index = ws.cell(row_idx, 8).value

            record = {
                "jichitai_code": str(jichitai_code).zfill(6),
                "prefecture_name": prefecture_name,
                "municipality_name": municipality_name,
                "finance": {
                    "financial_capability_index": normalize_value(financial_capability_index),  # 財政力指数
                    "current_balance_ratio": normalize_value(current_balance_ratio),  # 経常収支比率 (%)
                    "real_debt_service_ratio": normalize_value(real_debt_service_ratio),  # 実質公債費比率 (%)
                    "future_burden_ratio": normalize_value(future_burden_ratio),  # 将来負担比率 (%)
                    "laspeyres_index": normalize_value(laspeyres_index),  # ラスパイレス指数
                }
            }

            data.append(record)

        return data

    def get_by_code(self, jichitai_code: str) -> Optional[Dict]:
        """Get finance data for a specific municipality by code"""
        data = self.parse()
        code = str(jichitai_code).zfill(6)

        for record in data:
            if record["jichitai_code"] == code:
                return record

        return None

    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()