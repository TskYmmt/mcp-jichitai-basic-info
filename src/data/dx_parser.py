"""Parser for Digital Agency DX Dashboard data"""
from typing import Dict, List, Optional
from pathlib import Path
import openpyxl


class DXParser:
    """Parser for Digital Agency DX Dashboard data"""

    def __init__(self, comparison_file: str, online_procedures_file: str):
        self.comparison_file = Path(comparison_file)
        self.online_procedures_file = Path(online_procedures_file)
        self.wb_comparison = None
        self.wb_online = None

        if self.comparison_file.exists():
            self.wb_comparison = openpyxl.load_workbook(str(self.comparison_file), data_only=True)

        if self.online_procedures_file.exists():
            self.wb_online = openpyxl.load_workbook(str(self.online_procedures_file), data_only=True)

    def parse(self) -> List[Dict]:
        """
        Parse the DX Dashboard data files

        Returns:
            List of dictionaries containing municipality DX data
        """
        # まず比較データから市区町村一覧を取得
        municipalities = self._get_municipalities_from_comparison()

        # オンライン申請率データを追加
        self._add_online_procedures_data(municipalities)

        return list(municipalities.values())

    def _get_municipalities_from_comparison(self) -> Dict[str, Dict]:
        """
        Parse the comparison data file (横持ちデータ)

        Returns:
            Dictionary with municipality name as key
        """
        if not self.wb_comparison:
            return {}

        ws = self.wb_comparison.active
        municipalities = {}

        # 1行目に市区町村名（C列以降）
        for col_idx in range(3, ws.max_column + 1):
            municipality_name = ws.cell(1, col_idx).value
            if not municipality_name:
                continue

            municipality_name = str(municipality_name).strip()
            municipalities[municipality_name] = {
                "municipality": municipality_name,
                "dx_indicators": {}
            }

        # 2行目以降にDX指標データ
        for row_idx in range(2, ws.max_row + 1):
            category = ws.cell(row_idx, 1).value  # A列: カテゴリ
            indicator = ws.cell(row_idx, 2).value  # B列: 指標名

            if not indicator:
                continue

            indicator_key = str(indicator).strip()

            # 各市区町村のデータを取得
            for col_idx in range(3, ws.max_column + 1):
                municipality_name = ws.cell(1, col_idx).value
                if not municipality_name:
                    continue

                municipality_name = str(municipality_name).strip()
                value = ws.cell(row_idx, col_idx).value

                if municipality_name in municipalities:
                    municipalities[municipality_name]["dx_indicators"][indicator_key] = value

        return municipalities

    def _add_online_procedures_data(self, municipalities: Dict[str, Dict]):
        """
        Add online procedures data to municipalities

        Args:
            municipalities: Dictionary of municipalities to update
        """
        if not self.wb_online:
            return

        ws = self.wb_online.active

        # 1行目に市区町村名（C列以降）
        for col_idx in range(3, ws.max_column + 1):
            municipality_name = ws.cell(1, col_idx).value
            if not municipality_name:
                continue

            municipality_name = str(municipality_name).strip()

            if municipality_name not in municipalities:
                continue

            municipalities[municipality_name]["online_procedures"] = {}

            # 2行目以降に手続き名とオンライン申請率
            for row_idx in range(2, ws.max_row + 1):
                procedure_name = ws.cell(row_idx, 2).value  # B列: 手続き名

                if not procedure_name:
                    continue

                procedure_name = str(procedure_name).strip()
                value = ws.cell(row_idx, col_idx).value

                # パーセンテージ文字列を数値に変換（例: "88.8%" -> 88.8）
                if value and isinstance(value, str) and "%" in value:
                    try:
                        value = float(value.replace("%", "").strip())
                    except ValueError:
                        value = None

                municipalities[municipality_name]["online_procedures"][procedure_name] = value

    def get_by_name(self, municipality_name: str) -> Optional[Dict]:
        """
        Get DX data by municipality name

        Args:
            municipality_name: Municipality name

        Returns:
            Dictionary with DX data
        """
        all_data = self.parse()

        for record in all_data:
            if record["municipality"] == municipality_name:
                return record

        return None

    def close(self):
        """Close the workbooks"""
        if self.wb_comparison:
            self.wb_comparison.close()
        if self.wb_online:
            self.wb_online.close()